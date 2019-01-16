# -*- coding: utf-8 -*-

__author__ = 'ceho'

import os, sys
import traceback
import openpyxl
import datetime
import time

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import selenium.webdriver.support.ui as ui

userAgent = 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.99 Safari/537.36'

def findElement(node, css):
    try:
        res = node.find_element_by_css_selector(css)
    except NoSuchElementException:
        res = None
    except:
        res = None
    return res

def getText(node, css, attr = ''):
    res = ''
    elem = findElement(node, css)
    if elem:
        if attr == '':
            res = elem.text
        else:
            res = elem.get_attribute(attr)
        if not res:
            res = ''
    return res

def loginNarrpr():
    url = 'http://www.narrpr.com/'

    global driver
    driver.set_page_load_timeout(10)
    driver.get(url)

    count = 0
    while count < 30:
        count += 1
        time.sleep(1)
        elem = findElement(driver, 'input#SiteSearchForm_SearchTxt')
        if elem is not None:
            # already logged in
            return
        elem = findElement(driver, 'input[name="SignInEmail"]')
        if elem is not None:
            break

    if elem is None:
        return
    elem.clear()
    elem.send_keys(email)

    elem = findElement(driver, 'input[name="SignInPassword"]')
    elem.clear()
    elem.send_keys(password)
    elem.send_keys(Keys.RETURN)
    time.sleep(3)

def getPriceZillow(address):
    url = 'http://www.zillow.com/homes/"%s"_rb/' % address
    global driver
    try:
        driver.get(url)

        time.sleep(30)
        res = getText(driver, 'div.zest-value')
        if res == '':
            res = getText(driver, 'div[class="main-row  home-summary-row"] span')
    except TimeoutException:
        res = "timeout"
    except:
        res = "error: %s" % traceback.format_exc()
    return res

def getPriceRedfin(address):
    url = 'https://www.redfin.com/'
    global driver
    try:
        driver.get(url)
        wait = ui.WebDriverWait(driver, 30)
        wait.until(lambda driver: driver.current_url == url)

        elem = findElement(driver, 'input#search-box-input')
        elem.send_keys(address)
        elem.send_keys(Keys.RETURN)

        count = 0
        while count < 30:
            count += 1
            time.sleep(1)
            if driver.current_url != url:
                break
            elem = findElement(driver, 'div.noResultsView')
            if elem is not None:
                return "not found"
            elem = findElement(driver, 'div.resultsView')
            if elem is None:
                continue
            elem = findElement(elem, 'div.item-row a.item-title')
            if elem is None:
                continue
            driver.get(elem.get_attribute('href'))

        res = getText(driver, 'div.avmValue')
    except TimeoutException:
        res = "timeout"
    except:
        res = "error: %s" % traceback.format_exc()
    return res

def getPriceNarrpr(address):
    url = 'http://www.narrpr.com/'

    global driver
    try:
        driver.get(url)
        wait = ui.WebDriverWait(driver, 30)
        wait.until(lambda driver: driver.current_url == url)
    except:
        pass

    elem = findElement(driver, 'input#SiteSearchForm_SearchTxt')
    if elem is None:
        return "not found"
    elem.send_keys(address)
    elem = findElement(driver, 'a#SiteSearchForm_SearchBtn')
    elem.send_keys(Keys.RETURN)

    try:
        count = 0
        while count < 30:
            count += 1
            time.sleep(1)
            if driver.current_url != url:
                break
    except:
        pass

    res = getText(driver, 'div.priceSection div.price')
    return res

driver = None
ws = None

# for www.narrpr.com
email = 'mandiremax@gmail.com'
password = 'Schwartz424'

fname = 'Trustee Eval.xlsx'
sheet = 'Summary'
addrColumn = 3

zColumn = 9  # Zillow
rColumn = 10 # Redfin
nColumn = 8  # Narrpr

def main():
    chromeOptions = webdriver.ChromeOptions()
    chromeOptions.add_argument("user-agent=%s" % userAgent)
    chromeOptions.add_argument("user-data-dir=c:\ch")

    global driver
    driver = webdriver.Chrome(chrome_options=chromeOptions)
    wb = openpyxl.load_workbook(fname, data_only=True)

    global ws
    ws = wb[sheet]
    ws.row = 0
    rows = []
    data = []

    for row in range(2, ws.max_row+1):
        #if row > 11:
        #    break
        address = ws.cell(row=row, column=addrColumn).value
        if not address:
            continue
        address = address.replace('\n', ' ')
        #print address
        rows.append(row)
        data.append(address)

    try:
        # Zillow
        driver.set_page_load_timeout(15)
        for i, row in enumerate(rows):
            address = data[i]
            price = ws.cell(row=row, column=zColumn).value
            if price is None:
                price = getPriceZillow(address)
                ws.cell(row=row, column=zColumn, value=price)

        # Redfin
        driver.set_page_load_timeout(30)
        for i, row in enumerate(rows):
            address = data[i]
            price = ws.cell(row=row, column=rColumn).value
            if price is None:
                price = getPriceRedfin(address)
                ws.cell(row=row, column=rColumn, value=price)

        # Narrpr
        loginNarrpr()
        for i, row in enumerate(rows):
            address = data[i]
            price = ws.cell(row=row, column=nColumn).value
            if price is None:
                price = getPriceNarrpr(address)
                ws.cell(row=row, column=nColumn, value=price)
    finally:
        wb.save(fname)
        driver.close()

main()
